from fenrir import Blueprint, request, redirect, render_template, g
from routes.decorators import csrf_protect, login_required, admin_required, validate_object_id
from datetime import datetime, timezone
from bson import ObjectId
from models.student import get_students, create_student, update_student, delete_student
from models.class_model import get_classes, create_class, update_class, delete_class
from models.component import get_components, create_component, update_component, delete_component
from models.period import (
    get_academic_years, create_academic_year, toggle_academic_year_active, delete_academic_year,
    get_billing_periods, create_billing_period, update_billing_period, delete_billing_period,
    get_fee_configs, upsert_fee_config, delete_fee_config,
)
from models.audit import log_audit
import io

bp = Blueprint("master", url_prefix="/master")

# --- Siswa ---
@bp.get("/siswa")
@login_required
async def siswa_list():
    class_id = request.args.get("class_id", "")
    classes = await get_classes()
    students = await get_students(class_id if class_id else None)
    return render_template("master/siswa.html.j2", students=students, classes=classes, active_page="siswa")


@bp.post("/siswa")
@csrf_protect
@admin_required
@login_required
async def siswa_create():
    data = await request.form()
    try:
        await create_student(data)
        await log_audit(g.user["email"], "create", "students", None, None, data.get("name", ""))
    except ValueError as e:
        return render_template("master/siswa.html.j2", error=str(e),
                               students=await get_students(), classes=await get_classes(),
                               active_page="siswa")
    return redirect("/master/siswa")


@bp.post("/siswa/<id>/update")
@csrf_protect
@validate_object_id("id")
@admin_required
@login_required
async def siswa_update(id: str):
    data = await request.form()
    try:
        await update_student(id, data)
    except ValueError as e:
        return {"error": str(e)}, 400
    return redirect("/master/siswa")


@bp.post("/siswa/<id>/delete")
@csrf_protect
@validate_object_id("id")
@admin_required
@login_required
async def siswa_delete(id: str):
    await delete_student(id)
    return redirect("/master/siswa")


# --- Kelas ---
@bp.get("/kelas")
@login_required
async def kelas_list():
    ay_id = request.args.get("academic_year_id", "")
    academic_years = await get_academic_years()
    classes = await get_classes(ay_id if ay_id else None)
    return render_template("master/kelas.html.j2", classes=classes, academic_years=academic_years, active_page="kelas")


@bp.post("/kelas")
@csrf_protect
@admin_required
@login_required
async def kelas_create():
    data = await request.form()
    await create_class(data)  # using class_model's function
    return redirect("/master/kelas")


@bp.post("/kelas/<id>/update")
@csrf_protect
@validate_object_id("id")
@admin_required
@login_required
async def kelas_update(id: str):
    data = await request.form()
    await update_class(id, data)
    return redirect("/master/kelas")


@bp.post("/kelas/<id>/delete")
@csrf_protect
@validate_object_id("id")
@admin_required
@login_required
async def kelas_delete(id: str):
    await delete_class(id)
    return redirect("/master/kelas")


# --- Komponen ---
@bp.get("/komponen")
@login_required
async def komponen_list():
    components = await get_components()
    return render_template("master/komponen.html.j2", components=components, active_page="komponen")


@bp.post("/komponen")
@csrf_protect
@admin_required
@login_required
async def komponen_create():
    data = await request.form()
    await create_component(data)
    return redirect("/master/komponen")


@bp.post("/komponen/<id>/update")
@csrf_protect
@validate_object_id("id")
@admin_required
@login_required
async def komponen_update(id: str):
    data = await request.form()
    await update_component(id, data)
    return redirect("/master/komponen")


@bp.post("/komponen/<id>/delete")
@csrf_protect
@validate_object_id("id")
@admin_required
@login_required
async def komponen_delete(id: str):
    await delete_component(id)
    return redirect("/master/komponen")


# --- Periode ---
@bp.get("/periode")
@login_required
async def periode_list():
    academic_years = await get_academic_years()
    selected_ay_id = request.args.get("academic_year_id", "")
    periods = []
    fee_configs = []
    if selected_ay_id:
        periods = await get_billing_periods(selected_ay_id)
        fee_configs = await get_fee_configs(selected_ay_id)
    else:
        active_ay = next((ay for ay in academic_years if ay.get("is_active")), None)
        if active_ay:
            selected_ay_id = str(active_ay["_id"])
            periods = await get_billing_periods(selected_ay_id)
            fee_configs = await get_fee_configs(selected_ay_id)

    components = await get_components(active_only=True)
    selected_ay_name = ""
    for ay in academic_years:
        if str(ay["_id"]) == selected_ay_id:
            selected_ay_name = ay.get("name", "")
            break
    return render_template("master/periode.html.j2",
                           academic_years=academic_years,
                           selected_ay_id=selected_ay_id,
                           selected_ay_name=selected_ay_name,
                           periods=periods,
                           fee_configs=fee_configs,
                           components=components,
                           active_page="periode")


@bp.post("/periode/academic-year")
@csrf_protect
@admin_required
@login_required
async def academic_year_create():
    data = await request.form()
    await create_academic_year(data["name"])
    return redirect("/master/periode")


@bp.post("/periode/academic-year/<id>/activate")
@csrf_protect
@validate_object_id("id")
@admin_required
@login_required
async def academic_year_activate(id: str):
    await toggle_academic_year_active(id)
    return redirect("/master/periode")


@bp.post("/periode/academic-year/<id>/delete")
@csrf_protect
@validate_object_id("id")
@admin_required
@login_required
async def academic_year_delete(id: str):
    await delete_academic_year(id)
    return redirect("/master/periode")


@bp.post("/periode/billing-period")
@csrf_protect
@admin_required
@login_required
async def billing_period_create():
    data = await request.form()
    await create_billing_period(data)
    return redirect(f"/master/periode?academic_year_id={data['academic_year_id']}")


@bp.post("/periode/billing-period/<id>/update")
@csrf_protect
@validate_object_id("id")
@admin_required
@login_required
async def billing_period_update(id: str):
    data = await request.form()
    await update_billing_period(id, data)
    ay_id = data.get("academic_year_id", "")
    return redirect(f"/master/periode?academic_year_id={ay_id}")


@bp.post("/periode/billing-period/<id>/delete")
@csrf_protect
@validate_object_id("id")
@admin_required
@login_required
async def billing_period_delete(id: str):
    db = g.db
    period = await db.billing_periods.find_one({"_id": ObjectId(id)})
    await delete_billing_period(id)
    ay_id = str(period["academic_year_id"]) if period else ""
    return redirect(f"/master/periode?academic_year_id={ay_id}")


@bp.post("/periode/fee-config")
@csrf_protect
@admin_required
@login_required
async def fee_config_save():
    data = await request.form()
    await upsert_fee_config(data)
    return redirect(f"/master/periode?academic_year_id={data['academic_year_id']}")


@bp.post("/periode/fee-config/<id>/delete")
@csrf_protect
@validate_object_id("id")
@admin_required
@login_required
async def fee_config_delete(id: str):
    db = g.db
    config = await db.fee_configs.find_one({"_id": ObjectId(id)})
    await delete_fee_config(id)
    ay_id = str(config["academic_year_id"]) if config else ""
    return redirect(f"/master/periode?academic_year_id={ay_id}")


@bp.get("/kenaikan-kelas")
@login_required
@admin_required
async def kenaikan_kelas_page():
    db = g.db
    classes = await db.classes.find({}).sort("name", 1).to_list(None)
    return render_template("master/kenaikan-kelas.html.j2", classes=classes, active_page="kenaikan_kelas")

@bp.post("/kenaikan-kelas")
@login_required
@admin_required
@csrf_protect
async def kenaikan_kelas_action():
    db = g.db
    data = await request.form()
    from_class = data.get("from_class", "")
    to_class = data.get("to_class", "")
    to_angkatan = int(data.get("to_angkatan", 0))

    if not from_class or not to_class:
        return redirect("/master/kenaikan-kelas?error=invalid")

    from_class_doc = await db.classes.find_one({"_id": ObjectId(from_class)})
    to_class_doc = await db.classes.find_one({"_id": ObjectId(to_class)})

    if not from_class_doc or not to_class_doc:
        return redirect("/master/kenaikan-kelas?error=notfound")

    result = await db.students.update_many(
        {"class_id": ObjectId(from_class)},
        {"$set": {"class_id": ObjectId(to_class), "angkatan": to_angkatan, "updated_at": datetime.now(timezone.utc)}}
    )

    return redirect("/master/kenaikan-kelas?success=" + str(result.modified_count))


@bp.get("/siswa/import")
@login_required
@admin_required
async def import_siswa_page():
    return render_template("master/import-siswa.html.j2", active_page="import_siswa", result=None)


@bp.post("/siswa/import")
@login_required
@admin_required
@csrf_protect
async def import_siswa_action():
    from openpyxl import load_workbook
    db = g.db
    form_data = getattr(g, '_csrf_form_data', None) or await request.form()
    file = form_data.get("file")

    if not file or not hasattr(file, 'filename'):
        return render_template("master/import-siswa.html.j2", active_page="import_siswa", result={"error": "File tidak ditemukan"})

    try:
        content = await file.read()
        wb = load_workbook(io.BytesIO(content))
        ws = wb.active

        imported = 0
        skipped = 0
        errors = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            try:
                nis = str(row[0]).strip() if row[0] else ""
                name = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                class_name = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                angkatan = int(row[3]) if len(row) > 3 and row[3] else 0
                gender = str(row[4]).strip().upper() if len(row) > 4 and row[4] else "L"

                if not nis or not name:
                    skipped += 1
                    continue

                existing = await db.students.find_one({"nis": nis})
                if existing:
                    skipped += 1
                    continue

                class_doc = await db.classes.find_one({"name": class_name}) if class_name else None

                await db.students.insert_one({
                    "nis": nis,
                    "name": name,
                    "class_id": class_doc["_id"] if class_doc else None,
                    "angkatan": angkatan,
                    "gender": gender if gender in ("L", "P") else "L",
                    "created_at": datetime.now(timezone.utc),
                    "updated_at": datetime.now(timezone.utc),
                })
                imported += 1
            except Exception as e:  # noqa: BLE001
                errors.append(f"Baris {row[0]}: {e!s}")

        result = {"imported": imported, "skipped": skipped, "errors": errors}
        return render_template("master/import-siswa.html.j2", active_page="import_siswa", result=result)
    except Exception as e:
        return render_template("master/import-siswa.html.j2", active_page="import_siswa", result={"error": f"Gagal membaca file: {e!s}"})
