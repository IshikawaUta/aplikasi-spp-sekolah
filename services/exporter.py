import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def create_workbook(title: str) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = title
    return wb, ws

def set_header_style(ws, row: int, cols: int):
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font_w = Font(bold=True, size=11, color="FFFFFF")
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font_w
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

def auto_width(ws, cols: int):
    for col in range(1, cols + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = min(max_len + 4, 40)

def to_bytes(wb: Workbook) -> io.BytesIO:
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
